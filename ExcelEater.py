import openpyxl as xl

def Home():
  """
                  Home
  -------------------------------------
  1) Add a Master sheet
        - what is the name of the book
        - what is the name of the sheet
        - column(s) to match against child (index column)
  2) Add Child Sheet
        - what is the name of the book
        - what is the name of the sheet
        - column(s) to match against master (index column)
        - column(s) (values) to copy/paste into master
  3) Consume Child Sheet
        - overwrite master entries?
        - save updated master sheet
  4) Print Master values
  5) Print Child values
  6) Print values in Child missing from Master
  7) Print values in Master missing from Child
  8) Print values consumed by Master (and from which child)
  9) Add unmatched index values in child to master as new entries
        - which columns should we take values from in the child?
          - which columns do they match in the master?

  *** TYPE EXIT TO STOP THE PROGRAM *** 

  """

  print("                Home                ")
  print("-------------------------------------")
  print("1) Add a Master sheet")
  print("2) Add a Child sheet")
  print("3) Consume Child Sheet")
  print("4) Print Master values")
  print("5) Print Child values")
  print("-------------------------------------")
  print("*** TYPE EXIT TO STOP THE PROGRAM ***")


def GetValues(sheet, whichCol, index, start, end):
  """
  Return a nested dictionary of values from the target sheet
  create object of objects that contain the value to update as the key, and the values and xl cell as the details
  ie: {123456: {xlRow:144, xlCol:5,color: "blue", type:"sale", angle:90}, 890128:{xlRow:145, xlCol:5, color: "red", type:"return", angle:180}}
  """
  destDict = {}

  for i in range(start, end):
    indexVal = sheet.cell(i,index).value
    valueDict = {}
    for j in whichCol:
      valueDict["_row_"] = i
      valueDict[j] = sheet.cell(i,j).value
    
    destDict[indexVal] = valueDict
  return destDict


def ConsumeChild():
  pass

def GetInput(message, err="not a valid response"):
  """
  Get Input in a while loop
  """  
  stop = False
  while stop == False:
    print(message)
    try:
      userInput = input("> ")
      stop = True
    except:
      print(err)

  return userInput
  

def GetBook(prompt):
  """
  Return WorkBook
  """
  destination = False
  while destination == False:
    bookName = input(prompt)
    try:
      Book = xl.load_workbook(filename=bookName, read_only=False, keep_vba=True)
      destination = True
    except:
      print("Sorry, I couldn't find that filename in the current directory.  Did you forget the extension? xlsx, xlsm, etc...)> ")
  return Book, bookName


def GetSheet(book, prompt):
  """
  Return target sheet
  """
  worksheetID = False
  while worksheetID == False:
    worksheet = input(prompt)
    try:
      sheet = book[worksheet]
      worksheetID = True
    except: 
      print("Sorry, that sheet does not exist. Try again... ")
  return sheet


def ValidateMaxRow(sheet):
  """
  Did the user enter a valid ENDING cell?
  """
  valid = False
  while valid == False:
    maxRow_ = GetInput("Please enter the row you want to read the excel sheet to ", "Sorry, that is not a valid row. Please try again... ")
    if maxRow_ == "":
      print("Searching entire document...")
      maxRow_ = sheet.max_row +1
      valid = True
    try:
      maxRow_ = int(maxRow_)
      valid = True
    except:
      print("you must enter a valid number")
      continue
    if maxRow_ < 1:
      print("Searching entire document...")
      maxRow_ = sheet.max_row +1
      valid = True
  return maxRow_



def ValidateRowStart():
  """
  Did the user enter a valid STARTING cell?
  """
  valid = False
  while valid == False:
    rowStart_ = GetInput("Which row would you like to start on?", "Sorry that's not a valid row.  Please try again ")
    if rowStart_ == "":
      print("Starting at row 1...")
      rowStart_ = 1
      valid = True
    try:
      rowStart_ = int(rowStart_)
      valid = True
    except:
      print("you must enter a valid number")
      continue
    if rowStart_ <= 0:
      print("Starting at row 1...")
      rowStart_ = 0
      valid = True
  return rowStart_


def ValidateCol():
  """
  Did the user enter a VALID INDEX COLUMN?
  """
  valid = False
  while valid == False:
    colStart_ = GetInput("Which column are we going to use for the index?", "Sorry that's not a valid row.  Please try again ")
    if colStart_ == "":
      print("Using Column 1...")
      colStart_ = 1
      valid = True
    try:
      colStart_ = int(colStart_)
      valid = True
    except:
      print("you must enter a valid number")
      continue
    if colStart_ < 1:
      print("You did not specify a valid number. The program will guess at the which row is the START")
      colStart_ = 1
      valid = True
  return colStart_


def GetColumns(colCount):
  """
  Return an object containing the names and locations of the columns to retrieve data from
  """
  goodData = False
  whichCol = {}
  while goodData == False:
    for i in range(0, colCount):
      colName = GetInput("Name column " +str(i+1), "Invalid column name")
      colNum = GetInput("Enter Column number for value " +str(i+1), "Invalid column number")
      try:
        colNum = int(colNum)
        goodData = True
      except:
        print("Invalid column number...")
      whichCol[colNum] = colName
    
  return whichCol

def UploadSuccessful(sheet):
  if sheet == "master".lower():
    print("***************************************")
    print("*Master document uploaded successfully*")
    print("***************************************")
  if sheet == "child".lower():
    print("**************************************")
    print("*Child document uploaded successfully*")
    print("**************************************")

def PrintValues(values, msg="sheet hasn't been uploaded yet"):
  """
  Print the values in a more readable way
  """
  try:
    for v in values:
      print(str(v) + " - " + str(values[v]))
  except:
    print(msg)
  print("there are " + str(len(values)) + " values uploaded from this sheet")


def CheckColCount(masterColCount, childColCount):
  """
  Count the columns of Master, Count the columns of Child, compare the two, return which situation to address.
  """
  if masterColCount == childColCount:
    # straightAssignment()
    # columns are equal  Please map the columns:
      # for each column in child, assign a column in master that will receive the value
    return 1
  if masterColCount < childColCount:
    # master has too few columns.  Please merge child columns to fit.
      # please choose the first child column to merge
      # which column should it be merged with?
      # if mastercolCount == childColCount:
        #Assignment()
      # else:
        # 
    return 2
  if masterColCount > childColCount:
    # straightAssignment()
    # more columns in master than child  Please map the columns:
      # for each column in child, assign a column in master that will receive the value
    return 3
  
#master globals
masterBook = {}
masterSheet = {}
masterValues = {}
masterCol = {}
masterRowStart_ = 1
maxRow_ = 1
masterColCount = 0
masterBookName = ""

#child globals
childBook = {}
childSheet = {}
childDict = {}
childCol = {}
childRowStart_ = 1
childMaxRow_ = 1
childColCount = 0
missingCount = 0
childBookName_ = ""

#program loop
run = True
while run:
  Home()
  selection = str(GetInput("Please make a selection", "I'm sorry, that's not a valid selection.  Please try again"))
  print("you entered: " + selection)
  if selection.lower() == "exit":
    exit()
  if selection == "1":
    #get master sheet
    masterBookTuple = GetBook("What is the name of the MASTER book? ")
    masterBook = masterBookTuple[0]
    masterBookName = masterBookTuple[1]
    masterSheet = GetSheet(masterBook, "What is the name of the target sheet in the MASTER book? ")
    maxRow_ = ValidateMaxRow(masterSheet)
    masterRowStart_ = ValidateRowStart()
    masterColIndex = ValidateCol()
    masterColCount = GetInput("how many columns are we taking values from?","Sorry that's not a valid number.  Try again.")
    
    try:
      masterColCount = int(masterColCount)
    except:
      print("please enter a valid number")
      continue

    whichCol = GetColumns(masterColCount)

    masterValues = GetValues(masterSheet, whichCol, masterColIndex, masterRowStart_, maxRow_)
    UploadSuccessful(sheet="master")
  elif selection == "2":
    #get child sheet
    childBookTuple = GetBook("What is the name of the CHILD book? ")
    childBook_ = childBookTuple[0]
    childBookName_ = childBookTuple[1]
    childSheet_ = GetSheet(childBook_, "What is the name of the target sheet in the CHILD book? ")
    childMaxRow_ = ValidateMaxRow(childSheet_)
    childRowStart_ = ValidateRowStart()
    childColIndex = ValidateCol()
    childColCount = GetInput("How many columns are we taking values from?", "Sorry that's not a valid number. Try again.")

    try:
      childColCount = int(childColCount)
    except:
      print("please enter a valid number")
      continue

    childWhichCol = GetColumns(childColCount)

    childValues = GetValues(childSheet_, childWhichCol, childColIndex, childRowStart_, childMaxRow_)
    UploadSuccessful(sheet="child")
  elif selection == "3":
    # consume child sheet
    """
    TODO: overwrite values?
    overWrite = GetInput("overwrite values?")
    if overwrite == True:
      #overwrite values logic
    else:
      #only fill blanks in master logic
    """
    assignment = {}   

    for v in childWhichCol:
      absorbed = False
      while absorbed == False:
        print(whichCol)
        masterColAbsorb = GetInput("which master column will absorb '" + str(childWhichCol[v]) + "'")
        if masterColAbsorb.lower() == "exit":
          exit()
        try:
          masterColAbsorb = int(masterColAbsorb)
        except:
          print("please enter a valid number...")
        if masterColAbsorb in whichCol:
          absorbed = True
          assignment[v] = masterColAbsorb
          print("this is the assignment object " + str(assignment))
          # masterValues[masterColAbsorb] =
        else:
          print("That column doesn't exist in the master.  Try again...")


    optionSelected = False
    while optionSelected == False:
      overWrite = GetInput("PLEASE SELECT: (1) Overwrite values in Master  (2) Only fill blanks in Master ")
      if overWrite == "1":
        optionSelected = True
      elif overWrite == "2":
        optionSelected = True
      else:
        print("Please select option 1 or two...")
    
    totalColCount = CheckColCount(masterColCount, childColCount)
    if totalColCount == 1:
      foundCount = 0
      for v in childValues:
        if v in masterValues:
          foundCount += 1
          for e in masterValues[v]:
            if e != "_row_":
              # print(str(e) + " are the masterValue Keys")
              # print(str(masterValues[v][e]) + " are the masterValue values")
              # print(str(childValues[v][e]) + " are the child Values")
              masterSheet.cell(masterValues[v]['_row_'], e).value = childValues[v][e]
              print(str(childValues[v][e]))
        else:
          missingCount += 1
          print(str(v) + " is not in master")
      
      print(str(len(childValues)) + " total values in child")
      print(str(foundCount) + " matched")
      print(str(missingCount) + " unmatched in child")
      print(str(masterBookName) + " IS MASTER BOOK")
      masterBook.save(filename=masterBookName)
      exit()
      # print(masterBook)
      #masterbook.save(filename=masterBook)
  elif selection == "4":
    PrintValues(masterValues)
  elif selection == "5":
    PrintValues(childValues)


# get 4 col from master
# get 4 col from child
# assign which master col absorbs which child col
# if childColCount == masterColCount:
# 	straightAssignment()
# if childColCount < masterColCount:
# 	straightAssignment()
# if childColCount > masterColCount:
# 	select rows to merge
# 	assignment()


#CONSUME
# how many values be combined in the child?
  #if > 0, which letter ranges should be combined?
# which columns are we putting the values into in the master?

#OPTIONS
#ignore overwrite values in master from child?
  
#   wb.save(filename=dest)
#   print("saved updated workbook")
